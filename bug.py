import asyncio
import aiohttp
from lxml import etree
import openpyxl
from tkinter import *
timeout = aiohttp.ClientTimeout(total=5)

class get_url(object):
    def __init__(self,tx):
        #创建excel文件
        self.wb = openpyxl.Workbook()
        #tx是文本框,爬取过程中可以将爬取到的信息插入进文本框中,实现可视化
        self.tx = tx
        #清除文本框信息
        self.tx.delete('1.0', END)
        self.sheet = self.wb.active
        #信号量,控制协程数
        self.semaphore = asyncio.Semaphore(500)
        #UA伪装
        self.header = {
             "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.89 Safari/537.36"
        }
    #得到响应信息
    async def scrape(self, url):
        async with self.semaphore:
            session = aiohttp.ClientSession(headers=self.header)
            try:
                response = await session.get(url, timeout=timeout)
                result = await response.text()
                await session.close()
                return result
            except:
                return False
            finally:
                await session.close()
    #协程主体,在main方法中通过page值可以生成future列表,整合了scrape函数和parse函数
    async def scrape_index(self, page):
        url = f'https://www.zqystv.com/zqys/dianying/page/{page}.html/'
        #得到响应信息
        text = await self.scrape(url)
        #对信息进行处理
        await self.parse(text)
    #处理响应信息
    async def parse(self, text):
        try:
            tree = etree.HTML(text)
            #获得所有li标签
            li_list = tree.xpath('//div[@class="stui-pannel_bd"]/ul/li')
            for li in li_list:
                #获得url
                detail_url = li.xpath('./div/a/@href')[0]
                #插入到文本框中
                self.tx.insert(1.0,detail_url+'\n')
                #刷新防止卡死
                self.tx.update()
                #添加进文件中
                self.sheet.append([detail_url])
        except:
            pass
    def main(self):
        #一共1280页，可以创建一个协程列表
        scrape_index_tasks = [asyncio.ensure_future(self.scrape_index(page)) for page in range(1, 1280)]
        loop = asyncio.get_event_loop()
        tasks = asyncio.gather(*scrape_index_tasks)
        loop.run_until_complete(tasks)
        #保存到detail_url.xlsx文件
        self.wb.save('detail_url.xlsx')
        self.tx.delete("1.0", END)
        self.tx.insert(END, '\n\nurl爬取完毕,请打开文件detail_url.xlsx查看,进入下一步操作\n')

class analyse(object):
    def __init__(self,file,tx):
        self.wb = openpyxl.Workbook()
        self.tx = tx
        self.tx.delete('1.0',END)
        self.sheet = self.wb.active
        self.sheet.append(['电影名', '语言', '类型', '年份', '地区', '主演', '导演', '评分'])
        # 信号量，控制协程数
        self.semaphore = asyncio.Semaphore(500)
        self.movie_url = openpyxl.load_workbook(file).active
        self.max_row = self.movie_url.max_row
        self.header = {
             "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.89 Safari/537.36"
        }
    async def scrape(self, url):
        async with self.semaphore:
            session = aiohttp.ClientSession(headers=self.header)
            try:
                response = await session.get(url,timeout=timeout)
                result = await response.text()
                await session.close()
                return result
            except:
                return False
            finally:
                await session.close()
    async def get_url(self,row):
            return self.movie_url.cell(row=row,column=1).value

    async def scrape_index(self, row):
        url = await self.get_url(row)
        text = await self.scrape(url)
        await self.parse(text)

    async def parse(self, text):
        try:
            dict = {
                '电影名：':'',
                '语言：':'',
                '类型：':'',
                '年份：':'',
                '地区：':'',
                '主演：':'',
                '导演：':'',
                '评分：':''
            }
            detail_tree= etree.HTML(text)
            movie_title = detail_tree.xpath('//div[@class="stui-content__detail"]/h2/text()')
            score = detail_tree.xpath('//div[@class="stui-content__detail"]/div/span/text()')
            detail = detail_tree.xpath('//div[@class="stui-content__detail"]/p[1]//text()')
            actor = detail_tree.xpath('//div[@class="stui-content__detail"]/p[2]//text()')
            director = detail_tree.xpath('//div[@class="stui-content__detail"]/p[3]//text()')
            for i in range(0,len(detail)):
                if detail[i] in dict.keys():
                    j = i+1
                    lis = []
                    while detail[j] not in dict.keys() and detail[j]!='\xa0' and j<len(detail):
                        lis.append(detail[j])
                        j+=1
                    dict[detail[i]] = ' '.join(lis)

            if len(actor)>0:
                lis = []
                for i in actor:
                    if i in dict.keys() or i=='\xa0':
                        continue
                    else:
                        lis.append(i)
                dict['主演：'] = ' '.join(lis)
            if len(director)>0:
                lis = []
                for i in director:
                    if i in dict.keys() or i == '\xa0':
                        continue
                    else:
                        lis.append(i)
                dict['导演：'] = ' '.join(lis)
            dict['电影名：'] = movie_title[0].replace(" 剧情简介", "")
            dict['评分：'] = float(score[0])
            dict['年份：'] = int(dict['年份：'])
            str1 = ""
            str1 +="------------------------\n"
            for i,j in dict.items():
                str1 = str1+str(i)+str(j)+'\n'
            str1+="------------------------\n"
            self.tx.insert(1.0,str1)
            self.tx.update()
            lis = list(dict.values())
            self.sheet.append(lis)
        except:
            pass
    def main(self):
        scrape_index_tasks = [asyncio.ensure_future(self.scrape_index(row)) for row in range(1, self.max_row+1)]
        loop = asyncio.get_event_loop()
        tasks = asyncio.gather(*scrape_index_tasks)
        loop.run_until_complete(tasks)
        self.wb.save('movie.xlsx')
        self.tx.delete("1.0", END)
        self.tx.insert(END, '\n\n电影信息爬取完毕,请打开文件movie.xlsx查看,进入下一步操作\n')
